import pandas as pd
import json
from openpyxl import load_workbook
from PIL import Image
import fitz  # PyMuPDF
import subprocess
import io
import tempfile
import os
class ExcelDataProcessor:
    def extract_text_and_tables(self, file_path: str):
        """
        Extract both human-readable text and structured JSON from Excel:
        - Flattened sheet text for keyword detection
        - Structured JSON (sheets, charts, OCR results)
        """

        result = {}
        flat_text = ""

        # Step 1: Extract all sheet data with pandas
        try:
            xls = pd.read_excel(file_path, sheet_name=None)
            all_sheets = {}
            text_parts = []
            for sheet_name, df in xls.items():
                df = df.astype(str)
                records = df.to_dict(orient="records")
                all_sheets[sheet_name] = records
                # Flatten into text
                text_parts.append(f"--- {sheet_name} ---\n")
                for row in records:
                    text_parts.append(" ".join(row.values()))
            result["data"] = all_sheets
            flat_text = "\n".join(text_parts)
        except Exception as e:
            result["data_error"] = f"Failed to extract sheet data: {str(e)}"

        # Step 2: Detect charts & images
        try:
            wb = load_workbook(file_path)
            charts_info = {sheet.title: len(sheet._charts) for sheet in wb.worksheets}
            result["charts"] = charts_info

            ocr_results = {}
            for sheet in wb.worksheets:
                extracted_texts = []
                for img in getattr(sheet, "_images", []):
                    try:
                        img_bytes = img._data()
                        image = Image.open(io.BytesIO(img_bytes))

                        with tempfile.TemporaryDirectory() as tmp_dir:
                            pdf_path = os.path.join(tmp_dir, "img.pdf")
                            ocr_pdf_path = os.path.join(tmp_dir, "img_ocr.pdf")
                            image.convert("RGB").save(pdf_path)

                            subprocess.run(
                                ["ocrmypdf", "--force-ocr", pdf_path, ocr_pdf_path],
                                check=True, capture_output=True
                            )

                            doc = fitz.open(ocr_pdf_path)
                            ocr_text = "\n".join([page.get_text().strip() for page in doc])
                            doc.close()

                            if ocr_text.strip():
                                extracted_texts.append(ocr_text.strip())
                    except Exception as e:
                        extracted_texts.append(f"[OCR Error: {str(e)}]")

                ocr_results[sheet.title] = extracted_texts

            result["ocr_from_images"] = ocr_results
            flat_text += "\n" + "\n".join(
                ["\n".join(v) for v in ocr_results.values() if v]
            )
        except Exception as e:
            result["meta_error"] = f"Failed to detect charts/images: {str(e)}"

        return flat_text, result
    def extract_text_from_excel(self, file_path: str) -> str:
        """
        Extract maximum data from Excel:
        - All sheet data as text
        - Charts count per sheet
        - OCR text from embedded images using ocrmypdf
        """

        result = {}

        # Step 1: Extract all sheet data with pandas
        try:
            xls = pd.read_excel(file_path, sheet_name=None)
            all_sheets = {}
            for sheet_name, df in xls.items():
                df = df.astype(str)
                all_sheets[sheet_name] = df.to_dict(orient="records")
            result["data"] = all_sheets
        except Exception as e:
            result["data_error"] = f"Failed to extract sheet data: {str(e)}"

        # Step 2: Detect charts & images
        try:
            wb = load_workbook(file_path)
            charts_info = {sheet.title: len(sheet._charts) for sheet in wb.worksheets}
            result["charts"] = charts_info

            ocr_results = {}
            for sheet in wb.worksheets:
                extracted_texts = []
                for img in getattr(sheet, "_images", []):
                    try:
                        # Convert Excel image → PIL Image
                        img_bytes = img._data()
                        image = Image.open(io.BytesIO(img_bytes))

                        # Save image → temporary PDF
                        with tempfile.TemporaryDirectory() as tmp_dir:
                            img_path = os.path.join(tmp_dir, "img.png")
                            pdf_path = os.path.join(tmp_dir, "img.pdf")
                            ocr_pdf_path = os.path.join(tmp_dir, "img_ocr.pdf")

                            image.save(img_path)

                            # Convert image to PDF
                            image.convert("RGB").save(pdf_path)

                            # Run OCRmyPDF
                            subprocess.run(
                                ["ocrmypdf", "--force-ocr", pdf_path, ocr_pdf_path],
                                check=True, capture_output=True
                            )

                            # Extract OCR text from processed PDF
                            doc = fitz.open(ocr_pdf_path)
                            ocr_text = "\n".join([page.get_text().strip() for page in doc])
                            doc.close()

                            if ocr_text.strip():
                                extracted_texts.append(ocr_text.strip())

                    except Exception as e:
                        extracted_texts.append(f"[OCR Error: {str(e)}]")

                ocr_results[sheet.title] = extracted_texts

            result["ocr_from_images"] = ocr_results

        except Exception as e:
            result["meta_error"] = f"Failed to detect charts/images: {str(e)}"

        # Step 3: Return JSON string (safe for DB)
        return json.dumps(result, indent=2, ensure_ascii=False)
